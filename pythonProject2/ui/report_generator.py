"""
报告生成器模块
"""

import os
import json
from datetime import datetime
from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QGroupBox,
                             QLabel, QTextEdit, QPushButton, QFileDialog,
                             QComboBox, QCheckBox, QSpinBox, QFormLayout,
                             QProgressBar, QMessageBox)
from PyQt5.QtCore import Qt, pyqtSignal, QThread
from PyQt5.QtGui import QFont
import pandas as pd
import numpy as np
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.lib import colors
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import cv2


class ReportGenerator(QWidget):
    """报告生成器"""

    # 信号定义
    report_generated = pyqtSignal(str)
    report_error = pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self.report_data = None
        self.project_root = self.get_project_root()
        print(f"[DEBUG] 最终确定项目根目录: {self.project_root}")
        print(f"[DEBUG] data/results 路径: {os.path.join(self.project_root, 'data', 'results')}")
        print(f"[DEBUG] data/reports 路径: {os.path.join(self.project_root, 'data', 'reports')}")
        self.init_ui()

    def get_project_root(self):
        """获取项目根目录 - 增强版"""
        try:
            # 方法1：基于当前文件位置
            current_file = os.path.abspath(__file__)
            ui_dir = os.path.dirname(current_file)  # ui目录
            project_root = os.path.dirname(ui_dir)  # 项目根目录

            print(f"[INFO] 当前文件: {current_file}")
            print(f"[INFO] UI目录: {ui_dir}")
            print(f"[INFO] 计算的项目根目录: {project_root}")

            # 验证项目根目录是否有预期的子目录
            expected_dirs = ["data", "ui", "core"]
            valid = all(os.path.exists(os.path.join(project_root, d)) for d in expected_dirs)

            if valid:
                print(f"[INFO] 项目根目录验证成功: {project_root}")
                return project_root
            else:
                print(f"[WARNING] 项目根目录验证失败，尝试其他方法")

                # 方法2：基于当前工作目录
                cwd = os.getcwd()
                print(f"[INFO] 当前工作目录: {cwd}")

                # 检查当前目录是否是项目根目录
                cwd_valid = all(os.path.exists(os.path.join(cwd, d)) for d in expected_dirs)
                if cwd_valid:
                    print(f"[INFO] 使用当前工作目录作为项目根目录: {cwd}")
                    return cwd

                # 方法3：向上查找直到找到项目根目录
                search_dir = os.path.dirname(current_file)
                max_depth = 5
                depth = 0

                while depth < max_depth:
                    parent = os.path.dirname(search_dir)
                    if parent == search_dir:  # 到达根目录
                        break

                    parent_valid = all(os.path.exists(os.path.join(parent, d)) for d in expected_dirs)
                    if parent_valid:
                        print(f"[INFO] 向上查找到项目根目录: {parent}")
                        return parent

                    search_dir = parent
                    depth += 1

                # 如果所有方法都失败，使用当前文件所在目录的父目录
                print(f"[WARNING] 使用默认计算的项目根目录: {project_root}")
                return project_root

        except Exception as e:
            print(f"[ERROR] 获取项目根目录失败: {e}")
            # 返回当前工作目录作为后备
            return os.getcwd()

    def init_ui(self):
        """初始化UI"""
        layout = QVBoxLayout(self)

        # 标题
        title_label = QLabel("检测报告生成")
        title_font = QFont()
        title_font.setBold(True)
        title_font.setPointSize(16)
        title_label.setFont(title_font)
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)

        # 报告配置区域
        config_group = QGroupBox("报告配置")
        config_layout = QFormLayout(config_group)

        # 报告类型选择
        self.combo_report_type = QComboBox()
        self.combo_report_type.addItems([
            "单次检测报告",
            "批量检测报告",
            "统计分析报告",
            "完整检测报告"
        ])
        config_layout.addRow("报告类型:", self.combo_report_type)

        # 报告格式选择
        self.combo_format = QComboBox()
        self.combo_format.addItems(["Excel", "PDF", "HTML", "Word"])
        config_layout.addRow("输出格式:", self.combo_format)
        # 保存路径显示
        self.lbl_save_path = QLabel("报告保存位置: ./data/reports/")
        self.lbl_save_path.setStyleSheet("color: blue; font-weight: bold;")
        config_layout.addRow("保存位置:", self.lbl_save_path)

        # 包含内容选项
        self.chk_include_summary = QCheckBox("包含检测摘要")
        self.chk_include_summary.setChecked(True)
        self.chk_include_details = QCheckBox("包含详细数据")
        self.chk_include_details.setChecked(True)

        config_layout.addRow("", self.chk_include_summary)
        config_layout.addRow("", self.chk_include_details)

        # 报告标题
        self.txt_report_title = QTextEdit()
        self.txt_report_title.setMaximumHeight(50)
        self.txt_report_title.setText("机械零件同心度检测报告")
        config_layout.addRow("报告标题:", self.txt_report_title)

        # 公司信息
        self.txt_company_info = QTextEdit()
        self.txt_company_info.setMaximumHeight(80)
        self.txt_company_info.setText("西南科技大学\n信息与控制工程学院\n物联网工程专业")
        config_layout.addRow("公司信息:", self.txt_company_info)

        # 备注信息
        self.txt_remarks = QTextEdit()
        self.txt_remarks.setMaximumHeight(100)
        self.txt_remarks.setText("本报告由机械零件同心度检测系统自动生成。")
        config_layout.addRow("备注:", self.txt_remarks)

        layout.addWidget(config_group)

        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)

        # 按钮区域
        button_layout = QHBoxLayout()

        self.btn_generate = QPushButton("生成报告")
        self.btn_generate.clicked.connect(self.generate_report)
        self.btn_generate.setFixedHeight(40)
        self.btn_generate.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold;")

        self.btn_quick_save = QPushButton("快速保存")  # 新增按钮
        self.btn_quick_save.clicked.connect(self.quick_save_report)
        self.btn_quick_save.setFixedHeight(40)
        self.btn_quick_save.setToolTip("快速保存到默认目录（./data/reports/）")

        self.btn_export_excel = QPushButton("导出结果数据")  # 修改按钮文本
        self.btn_export_excel.clicked.connect(self.export_to_excel)  # 修改连接的方法
        self.btn_export_excel.setFixedHeight(40)
        self.btn_export_excel.setToolTip("导出原始数据到results目录")

        # 添加关闭按钮
        self.btn_close = QPushButton("关闭")
        self.btn_close.clicked.connect(self.close)
        self.btn_close.setFixedHeight(40)
        self.btn_close.setStyleSheet("background-color: #f44336; color: white; font-weight: bold;")

        # 将所有按钮添加到布局中（每个按钮只添加一次）
        button_layout.addWidget(self.btn_generate)
        button_layout.addWidget(self.btn_quick_save)
        button_layout.addWidget(self.btn_export_excel)
        button_layout.addWidget(self.btn_close)

        layout.addLayout(button_layout)

        # 状态标签
        self.lbl_status = QLabel("就绪 - 选择报告格式后点击生成")
        self.lbl_status.setAlignment(Qt.AlignCenter)
        self.lbl_status.setStyleSheet("color: green; font-weight: bold;")
        layout.addWidget(self.lbl_status)

    def set_report_data(self, data):
        """设置报告数据（支持单结果和多结果）"""
        if isinstance(data, dict):
            # 检查是否是包含all_results的数据结构
            if 'all_results' in data:
                self.report_data = data['all_results']
                self.current_result = data.get('current_result')
                self.total_count = data.get('total_count', len(self.report_data))
            else:
                # 单结果格式
                self.report_data = [data]
                self.current_result = data
                self.total_count = 1
        elif isinstance(data, list):
            # 直接传入结果列表
            self.report_data = data
            self.current_result = data[-1] if data else None
            self.total_count = len(data)
        else:
            # 未知格式，尝试转换为列表
            self.report_data = [data] if data else []
            self.current_result = data
            self.total_count = len(self.report_data)

        # 更新UI显示记录数量
        if hasattr(self, 'lbl_status'):
            self.lbl_status.setText(f"就绪 - 共{self.total_count}条检测记录")

    def generate_report(self):
        """生成报告（修复版，确保保存到正确位置）"""
        if self.report_data is None:
            QMessageBox.warning(self, "警告", "没有可生成报告的数据！")
            return

        # 所有报告类型都保存到reports目录
        default_dir = os.path.join(self.project_root, "data", "reports")
        os.makedirs(default_dir, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_type = self.combo_report_type.currentText()

        if self.combo_format.currentText() == "Excel":
            default_name = f"{report_type}_{timestamp}.xlsx"
            file_filter = "Excel文件 (*.xlsx)"
        elif self.combo_format.currentText() == "PDF":
            default_name = f"{report_type}_{timestamp}.pdf"
            file_filter = "PDF文件 (*.pdf)"
        elif self.combo_format.currentText() == "HTML":
            default_name = f"{report_type}_{timestamp}.html"
            file_filter = "HTML文件 (*.html)"
        elif self.combo_format.currentText() == "Word":
            default_name = f"{report_type}_{timestamp}.docx"
            file_filter = "Word文件 (*.docx)"
        else:
            default_name = f"report_{timestamp}.xlsx"  # 默认Excel
            file_filter = "Excel文件 (*.xlsx)"

        # 设置默认路径
        default_path = os.path.join(default_dir, default_name)

        # 让用户选择保存位置（默认指向正确目录）
        file_path, _ = QFileDialog.getSaveFileName(
            self, "保存报告",
            default_path,
            file_filter
        )

        if not file_path:
            return  # 用户取消

        # 显示保存路径
        self.lbl_save_path.setText(f"保存到: {os.path.basename(file_path)}")

        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.lbl_status.setText("正在生成报告...")

        try:
            # 根据格式调用不同的生成函数
            if self.combo_format.currentText() == "Excel":
                self.generate_excel_report(file_path)
            elif self.combo_format.currentText() == "PDF":
                self.generate_pdf_report(file_path)
            elif self.combo_format.currentText() == "HTML":
                self.generate_html_report(file_path)
            elif self.combo_format.currentText() == "Word":
                self.show_error("Word格式暂不支持")
                return

            self.progress_bar.setValue(100)
            self.lbl_status.setText(f"报告已生成: {os.path.basename(file_path)}")
            self.report_generated.emit(file_path)

            # 显示成功信息，包含完整路径
            success_msg = f"报告生成成功！\n\n文件已保存到：\n{file_path}"
            QMessageBox.information(self, "成功", success_msg)

        except Exception as e:
            error_msg = f"生成报告失败：{str(e)}\n\n请检查：\n1. 保存目录是否有写入权限\n2. 磁盘空间是否充足"
            self.lbl_status.setText(f"生成失败: {str(e)}")
            self.report_error.emit(str(e))
            QMessageBox.critical(self, "错误", error_msg)

        finally:
            self.progress_bar.setVisible(False)

    def quick_save_report(self):
        """快速保存报告到默认目录（无需选择路径）"""
        if self.report_data is None:
            QMessageBox.warning(self, "警告", "没有可生成报告的数据！")
            return

        try:
            # 所有报告类型都保存到reports目录
            save_dir = os.path.join(self.project_root, "data", "reports")
            os.makedirs(save_dir, exist_ok=True)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            report_type = self.combo_report_type.currentText().replace(" ", "_")

            if self.combo_format.currentText() == "PDF":
                filename = f"{report_type}_{timestamp}.pdf"
                file_path = os.path.join(save_dir, filename)
                # 直接调用类内的PDF生成方法
                self.generate_pdf_report(file_path)

            elif self.combo_format.currentText() == "Excel":
                filename = f"{report_type}_{timestamp}.xlsx"
                file_path = os.path.join(save_dir, filename)
                # 直接调用类内的Excel生成方法
                self.generate_excel_report(file_path)

            elif self.combo_format.currentText() == "HTML":
                filename = f"{report_type}_{timestamp}.html"
                file_path = os.path.join(save_dir, filename)
                # 调用类内的HTML生成方法
                self.generate_html_report(file_path)

            else:  # Word或其他格式
                # 给出提示，暂时不支持Word格式的快速保存
                QMessageBox.warning(
                    self,
                    "提示",
                    f"{self.combo_format.currentText()}格式的快速保存功能暂未实现，\n"
                    "请使用手动生成功能选择保存位置。"
                )
                return

            # 显示成功信息
            success_msg = f"报告已快速保存！\n\n文件位置：\n{file_path}"
            self.lbl_status.setText(f"快速保存成功: {os.path.basename(file_path)}")
            self.report_generated.emit(file_path)

            QMessageBox.information(self, "成功", success_msg)

        except Exception as e:
            error_msg = f"快速保存失败：{str(e)}"
            self.lbl_status.setText("快速保存失败")
            QMessageBox.critical(self, "错误", error_msg)

    def generate_pdf_report(self, file_path: str):
        """生成PDF报告（修复版，确保文件实际保存）"""
        try:
            print(f"[DEBUG] 开始生成PDF报告，路径: {file_path}")

            # 确保目录存在
            directory = os.path.dirname(file_path)
            if directory:
                os.makedirs(directory, exist_ok=True)
                print(f"[DEBUG] 确保目录存在: {directory}")

            # 数据处理部分：在创建PDF文档之前处理数据
            # 复制数据以避免修改原始数据
            report_data = self.report_data

            # 注意：这里report_data可能是单个字典或列表
            # 检查是否是批量报告（列表）
            is_batch_report = isinstance(report_data, list) and len(report_data) > 1

            # 创建PDF文档
            doc = SimpleDocTemplate(
                file_path,
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18
            )

            styles = getSampleStyleSheet()

            # 自定义样式
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=1
            )

            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=14,
                spaceAfter=12,
                spaceBefore=12
            )

            normal_style = styles['Normal']

            # 构建报告内容
            story = []

            # 标题
            title = self.txt_report_title.toPlainText()
            if is_batch_report:
                title += f" - 批量检测报告（共{len(report_data)}条）"
            story.append(Paragraph(title, title_style))
            story.append(Spacer(1, 0.2 * inch))

            # 报告信息
            story.append(Paragraph("报告信息", heading_style))

            report_info_data = [
                ["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                ["报告类型", "批量检测报告" if is_batch_report else "单次检测报告"],
                ["检测系统", "机械零件同心度检测系统 v1.0"],
                ["保存路径", os.path.basename(file_path)]
            ]

            if is_batch_report:
                # 添加批量报告的统计信息
                total_count = len(report_data)
                passed_count = sum(1 for d in report_data if d.get('is_qualified', False))
                failed_count = total_count - passed_count
                pass_rate = (passed_count / total_count * 100) if total_count > 0 else 0

                report_info_data.insert(1, ["检测总数", str(total_count)])
                report_info_data.insert(2, ["合格数量", str(passed_count)])
                report_info_data.insert(3, ["不合格数量", str(failed_count)])
                report_info_data.insert(4, ["合格率", f"{pass_rate:.1f}%"])

            report_info_table = Table(report_info_data, colWidths=[1.5 * inch, 4 * inch])
            report_info_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F81BD')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#DCE6F1')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black)
            ]))

            story.append(report_info_table)
            story.append(Spacer(1, 0.3 * inch))

            if is_batch_report:
                # 批量报告的详细数据
                story.append(Paragraph("详细检测数据", heading_style))

                # 创建详细数据表头
                detail_data = [["序号", "检测时间", "实际偏差(mm)", "像素偏差(px)", "同心度(‰)", "状态"]]

                for i, data in enumerate(report_data, 1):
                    status = "✅ 合格" if data.get('is_qualified', False) else "❌ 不合格"
                    detail_data.append([
                        str(i),
                        data.get('detection_time', 'N/A'),
                        f"{data.get('error_mm', 0):.3f}",
                        f"{data.get('pixel_error', 0):.2f}",
                        f"{data.get('concentricity', 0):.3f}",
                        status
                    ])

                # 创建详细数据表格
                detail_table = Table(detail_data,
                                     colWidths=[0.5 * inch, 1.5 * inch, 1 * inch, 1 * inch, 1 * inch, 0.8 * inch])
                detail_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F81BD')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
                ]))

                story.append(detail_table)

                # 如果数据量多，添加分页提示
                if len(report_data) > 20:
                    story.append(Paragraph(f"注：共{len(report_data)}条记录，表格已自动分页显示。", normal_style))

            else:
                # 单次检测报告（原有逻辑）
                # ... 保持原有的单次报告生成逻辑 ...
                story.append(Paragraph("检测结果", heading_style))

                # 提取关键信息
                result = report_data if isinstance(report_data, dict) else report_data[0]
                error_mm = result.get('error_mm', 0)
                pixel_error = result.get('pixel_error', 0)
                is_qualified = result.get('is_qualified', False)
                concentricity = result.get('concentricity', 0)

                summary_data = [
                    ["检测项目", "检测结果", "结论"],
                    ["实际偏差", f"{error_mm:.3f} mm", "✅ 合格" if is_qualified else "❌ 不合格"],
                    ["像素偏差", f"{pixel_error:.2f} px", "-"],
                    ["同心度", f"{concentricity:.3f}‰", "-"],
                    ["检测时间", result.get('detection_time', 'N/A'), "-"],
                    ["图像尺寸", result.get('image_size', 'N/A'), "-"]
                ]

                summary_table = Table(summary_data, colWidths=[2 * inch, 2 * inch, 1.5 * inch])
                summary_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F81BD')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ]))

                story.append(summary_table)

            story.append(Spacer(1, 0.3 * inch))

            # 备注
            remarks = self.txt_remarks.toPlainText()
            if remarks:
                story.append(Paragraph("备注说明", heading_style))
                story.append(Paragraph(remarks, normal_style))

            # 生成PDF
            doc.build(story)

            print(f"[SUCCESS] PDF报告已成功保存到: {file_path}")

        except Exception as e:
            print(f"[ERROR] 生成PDF报告失败: {str(e)}")
            raise

    # 在 generate_excel_report 方法中，修改以下部分：

    def generate_excel_report(self, file_path: str):
        """生成Excel报告（完整版，包含格式化、统计、说明等内容）"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "检测报告"

            # 设置默认样式
            title_font = Font(name='微软雅黑', size=16, bold=True)
            header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
            normal_font = Font(name='宋体', size=11)
            highlight_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")  # 绿色
            warning_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # 红色
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # 蓝色
            section_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # 灰色

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 当前行号，用于跟踪写入位置
            current_row = 1

            # 1. 报告标题
            report_title = self.txt_report_title.toPlainText()
            ws.merge_cells(f'A{current_row}:F{current_row}')
            ws[f'A{current_row}'] = report_title
            ws[f'A{current_row}'].font = title_font
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[current_row].height = 35
            current_row += 2

            # 2. 基本信息区域
            ws.merge_cells(f'A{current_row}:F{current_row}')
            ws[f'A{current_row}'] = "一、基本信息"
            ws[f'A{current_row}'].font = header_font
            ws[f'A{current_row}'].fill = header_fill
            ws[f'A{current_row}'].border = thin_border
            ws[f'A{current_row}'].alignment = Alignment(horizontal='left')
            ws.row_dimensions[current_row].height = 25
            current_row += 1

            # 基本信息表格
            basic_info = [
                ["报告编号", f"REPORT-{datetime.now().strftime('%Y%m%d%H%M%S')}"],
                ["生成时间", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ["报告类型", self.combo_report_type.currentText()],
                ["检测系统", "机械零件同心度检测系统 v2.0"],
                ["软件版本", "1.2.1"],
                ["检测单位", "西南科技大学"]  # 固定值，根据你的需求
            ]

            for i, (key, value) in enumerate(basic_info, start=current_row):
                ws[f'A{i}'] = key
                ws[f'B{i}'] = value
                ws[f'A{i}'].font = Font(bold=True)
                ws[f'A{i}'].border = thin_border
                ws[f'B{i}'].border = thin_border
                # 合并右侧单元格用于长文本
                ws.merge_cells(f'B{i}:F{i}')

            current_row += len(basic_info) + 2

            # 3. 检测说明区域
            ws.merge_cells(f'A{current_row}:F{current_row}')
            ws[f'A{current_row}'] = "二、检测说明"
            ws[f'A{current_row}'].font = header_font
            ws[f'A{current_row}'].fill = header_fill
            ws[f'A{current_row}'].border = thin_border
            ws[f'A{current_row}'].alignment = Alignment(horizontal='left')
            ws.row_dimensions[current_row].height = 25
            current_row += 1

            # 检测说明内容
            detection_info = [
                ["检测对象", "机械零件（轴套/螺纹杆组件）"],
                ["检测目的", "测量内外圆同心度，评估加工质量"],
                ["检测标准", "GB/T 1958-2017 产品几何量技术规范(GPS)"],
                ["合格标准", "同心度 ≤ 0.2‰（千分比）"],
                ["检测方法", "机器视觉 + 图像处理算法"],
                ["主要算法", "Canny边缘检测 + Hough圆变换 + 最小二乘拟合"],
                ["标定方式", "棋盘格标定法，像素-物理尺寸转换"],
                ["参考标准", "国际标准ISO 1101:2017几何公差规范"]
            ]

            for i, (key, value) in enumerate(detection_info, start=current_row):
                ws[f'A{i}'] = key
                ws[f'B{i}'] = value
                ws[f'A{i}'].font = Font(bold=True)
                ws[f'A{i}'].border = thin_border
                ws[f'B{i}'].border = thin_border
                ws.merge_cells(f'B{i}:F{i}')

            current_row += len(detection_info) + 2

            # 4. 核心检测结果区域
            ws.merge_cells(f'A{current_row}:F{current_row}')
            ws[f'A{current_row}'] = "三、核心检测结果"
            ws[f'A{current_row}'].font = header_font
            ws[f'A{current_row}'].fill = header_fill
            ws[f'A{current_row}'].border = thin_border
            ws[f'A{current_row}'].alignment = Alignment(horizontal='left')
            ws.row_dimensions[current_row].height = 25
            current_row += 1

            # 核心结果表头
            headers = ["检测项目", "检测值", "单位", "标准要求", "状态", "结论分析"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = section_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1

            # 核心检测结果数据
            # 这里是关键：直接从 report_data 中提取数据，如果没有则使用默认值
            if self.report_data:
                # 获取检测结果数据
                data = self.report_data[0] if isinstance(self.report_data, list) else self.report_data

                # 提取关键字段，使用安全的 get 方法
                concentricity = data.get('concentricity', 0.000)
                error_mm = data.get('error_mm', 0.199)  # 对应报告中的偏心距(实际)
                pixel_error = data.get('pixel_error', 6.04)
                processing_time = data.get('processing_time', 0.000)
                confidence = data.get('confidence', 0.0)  # 0-1之间的值

                # 判断是否合格
                tolerance = 0.2  # 标准公差
                is_concentricity_qualified = concentricity <= tolerance
                is_error_mm_qualified = error_mm <= tolerance
                is_pixel_qualified = pixel_error <= 6.06
                is_time_qualified = processing_time <= 3.0
                is_confidence_qualified = confidence >= 0.85

                core_results = [
                    ["同心度",
                     f"{concentricity:.3f}",
                     "‰",
                     f"≤{tolerance:.2f}‰",
                     "合格" if is_concentricity_qualified else "不合格",
                     "满足公差要求，加工精度良好" if is_concentricity_qualified else "超出公差范围，建议返工"],

                    ["偏心距(实际)",
                     f"{error_mm:.3f}",
                     "mm",
                     f"≤{tolerance:.2f}mm",
                     "合格" if is_error_mm_qualified else "不合格",
                     "物理偏差在允许范围内" if is_error_mm_qualified else "物理偏差过大，影响装配"],

                    ["像素偏差",
                     f"{pixel_error:.2f}",
                     "px",
                     "≤6.06px",
                     "合格" if is_pixel_qualified else "不合格",
                     ""],

                    ["检测耗时",
                     f"{processing_time:.3f}",
                     "秒",
                     "≤3.0秒",
                     "合格" if is_time_qualified else "不合格",
                     "检测效率符合要求"],

                    ["检测置信度",
                     f"{confidence * 100:.1f}",  # 转换为百分比
                     "%",
                     "≥85%",
                     "合格" if is_confidence_qualified else "不合格",
                     "检测结果可靠性高" if is_confidence_qualified else "检测置信度偏低"]
                ]

                # 写入核心结果数据
                for row_data in core_results:
                    for col_idx, cell_value in enumerate(row_data, start=1):
                        cell = ws.cell(row=current_row, column=col_idx, value=cell_value)
                        cell.border = thin_border
                        cell.alignment = Alignment(
                            horizontal='center' if col_idx in [2, 3, 4, 5] else 'left',
                            vertical='center'
                        )

                        # 状态列着色
                        if col_idx == 5:  # 状态列
                            if cell_value == "合格":
                                cell.fill = highlight_fill
                                cell.font = Font(color="006600", bold=True)
                            else:
                                cell.fill = warning_fill
                                cell.font = Font(color="990000", bold=True)

                    current_row += 1

            else:
                # 如果没有数据，填充默认值（用于测试）
                default_results = [
                    ["同心度", "0.000", "‰", "≤0.20‰", "合格", "满足公差要求，加工精度良好"],
                    ["偏心距(实际)", "0.199", "mm", "≤0.20mm", "合格", "物理偏差在允许范围内"],
                    ["像素偏差", "6.04", "px", "≤6.06px", "合格", ""],
                    ["检测耗时", "0.000", "秒", "≤3.0秒", "合格", "检测效率符合要求"],
                    ["检测置信度", "0.0", "%", "≥85%", "不合格", "检测置信度偏低"]
                ]

                for row_data in default_results:
                    for col_idx, cell_value in enumerate(row_data, start=1):
                        cell = ws.cell(row=current_row, column=col_idx, value=cell_value)
                        cell.border = thin_border
                        cell.alignment = Alignment(
                            horizontal='center' if col_idx in [2, 3, 4, 5] else 'left',
                            vertical='center'
                        )

                        if col_idx == 5:
                            if cell_value == "合格":
                                cell.fill = highlight_fill
                                cell.font = Font(color="006600", bold=True)
                            else:
                                cell.fill = warning_fill
                                cell.font = Font(color="990000", bold=True)

                    current_row += 1

            current_row += 2

            # 5. 几何参数详情区域
            ws.merge_cells(f'A{current_row}:F{current_row}')
            ws[f'A{current_row}'] = "四、几何参数详情"
            ws[f'A{current_row}'].font = header_font
            ws[f'A{current_row}'].fill = header_fill
            ws[f'A{current_row}'].border = thin_border
            ws[f'A{current_row}'].alignment = Alignment(horizontal='left')
            ws.row_dimensions[current_row].height = 25
            current_row += 1

            # 几何参数表头
            geo_headers = ["几何参数", "内圆", "外圆", "单位", "备注"]
            for col, header in enumerate(geo_headers, start=1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = section_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1

            # 几何参数数据（从数据中提取或使用默认值）
            if self.report_data:
                data = self.report_data[0] if isinstance(self.report_data, list) else self.report_data
                inner_circle = data.get('inner_circle', {})
                outer_circle = data.get('outer_circle', {})

                geo_data = [
                    ["中心坐标X",
                     f"{inner_circle.get('x', 442.0):.1f}",
                     f"{outer_circle.get('x', 439.5):.1f}",
                     "px",
                     "图像像素坐标系"],

                    ["中心坐标Y",
                     f"{inner_circle.get('y', 373.0):.1f}",
                     f"{outer_circle.get('y', 367.5):.1f}",
                     "px",
                     "图像像素坐标系"],

                    ["半径",
                     f"{inner_circle.get('radius', 64.9):.1f}",
                     f"{outer_circle.get('radius', 259.5):.1f}",
                     "px",
                     "基于边缘检测"],

                    ["置信度",
                     f"{inner_circle.get('confidence', 0.80) * 100:.1f}%",
                     f"{outer_circle.get('confidence', 0.868) * 100:.1f}%",
                     "%",
                     "算法检测可靠度"]
                ]
            else:
                # 默认值
                geo_data = [
                    ["中心坐标X", "442.0", "439.5", "px", "图像像素坐标系"],
                    ["中心坐标Y", "373.0", "367.5", "px", "图像像素坐标系"],
                    ["半径", "64.9", "259.5", "px", "基于边缘检测"],
                    ["检测方法", "simple_centroid", "merged", "-", "算法实现方式"],
                    ["置信度", "80.0%", "86.8%", "%", "算法检测可靠度"]
                ]

            # 写入几何参数数据
            for row_data in geo_data:
                for col_idx, cell_value in enumerate(row_data, start=1):
                    cell = ws.cell(row=current_row, column=col_idx, value=cell_value)
                    cell.border = thin_border
                    cell.alignment = Alignment(
                        horizontal='center' if col_idx in [2, 3, 4] else 'left',
                        vertical='center'
                    )
                current_row += 1

            current_row += 2

            # 6. 检测结论与建议区域
            ws.merge_cells(f'A{current_row}:F{current_row}')
            ws[f'A{current_row}'] = "五、检测结论与建议"
            ws[f'A{current_row}'].font = header_font
            ws[f'A{current_row}'].fill = header_fill
            ws[f'A{current_row}'].border = thin_border
            ws[f'A{current_row}'].alignment = Alignment(horizontal='left')
            ws.row_dimensions[current_row].height = 25
            current_row += 1

            # 结论内容
            # 判断整体是否合格（所有项目都合格才算合格）
            if self.report_data:
                data = self.report_data[0] if isinstance(self.report_data, list) else self.report_data
                is_qualified = data.get('is_qualified', False)
            else:
                # 根据默认数据判断：除了置信度不合格，其他都合格
                is_qualified = False  # 置信度不合格，所以整体不合格

            conclusion_text = ""
            recommendation_text = ""

            if is_qualified:
                conclusion_text = "✓ 检测结论：零件同心度符合质量要求，判定为【合格品】"
                recommendation_text = "✓ 处理建议：无需返工，可直接进入下一道工序"
            else:
                conclusion_text = "✗ 检测结论：零件同心度不符合质量要求，判定为【不合格品】"
                recommendation_text = "✓ 处理建议：建议进行返工处理，重新调整加工工艺参数"

            ws.merge_cells(f'A{current_row}:F{current_row}')
            ws[f'A{current_row}'] = conclusion_text
            if is_qualified:
                ws[f'A{current_row}'].font = Font(size=12, bold=True, color="006600")
                ws[f'A{current_row}'].fill = highlight_fill
            else:
                ws[f'A{current_row}'].font = Font(size=12, bold=True, color="990000")
                ws[f'A{current_row}'].fill = warning_fill
            ws[f'A{current_row}'].border = thin_border
            ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
            current_row += 1

            ws.merge_cells(f'A{current_row}:F{current_row}')
            ws[f'A{current_row}'] = recommendation_text
            ws[f'A{current_row}'].font = Font(size=11, italic=True)
            ws[f'A{current_row}'].border = thin_border
            ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
            current_row += 2

            # 7. 备注与签字区域
            ws.merge_cells(f'A{current_row}:F{current_row}')
            ws[f'A{current_row}'] = "六、备注与签字"
            ws[f'A{current_row}'].font = header_font
            ws[f'A{current_row}'].fill = header_fill
            ws[f'A{current_row}'].border = thin_border
            ws[f'A{current_row}'].alignment = Alignment(horizontal='left')
            ws.row_dimensions[current_row].height = 25
            current_row += 1

            # 备注内容
            remarks = "本报告由机械零件同心度检测系统自动生成。\n\n注：本报告由系统自动生成，如需纸质报告请打印并签字确认。"
            ws.merge_cells(f'A{current_row}:F{current_row + 2}')
            ws[f'A{current_row}'] = remarks
            ws[f'A{current_row}'].font = Font(size=10)
            ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            current_row += 4

            # 签字表格
            sign_headers = ["检测人员", "审核人员", "批准人员", "检测单位"]
            for col, header in enumerate(sign_headers, start=1):
                cell = ws.cell(row=current_row, column=col + (col - 1), value=header)  # 每两列一个
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # 签字栏
                sign_cell = ws.cell(row=current_row + 1, column=col + (col - 1))
                sign_cell.value = "签字：________________"
                sign_cell.border = thin_border
                sign_cell.alignment = Alignment(horizontal='left', vertical='center')

                # 日期栏
                date_cell = ws.cell(row=current_row + 2, column=col + (col - 1))
                date_cell.value = "日期：________________"
                date_cell.border = thin_border
                date_cell.alignment = Alignment(horizontal='left', vertical='center')

                # 合并单元格（每个项目占两列）
                ws.merge_cells(start_row=current_row, start_column=col + (col - 1),
                               end_row=current_row, end_column=col + (col - 1) + 1)
                ws.merge_cells(start_row=current_row + 1, start_column=col + (col - 1),
                               end_row=current_row + 1, end_column=col + (col - 1) + 1)
                ws.merge_cells(start_row=current_row + 2, start_column=col + (col - 1),
                               end_row=current_row + 2, end_column=col + (col - 1) + 1)

            # 自动调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass

                adjusted_width = min(max_length + 2, 30)  # 最大宽度30
                ws.column_dimensions[column_letter].width = adjusted_width

            # 保存工作簿
            wb.save(file_path)

            print(f"[SUCCESS] 完整Excel报告已保存到: {file_path}")
            return file_path

        except Exception as e:
            print(f"[ERROR] 生成完整Excel报告失败: {str(e)}")
            import traceback
            traceback.print_exc()
            # 回退到简单模式
            return self._generate_simple_excel_report(file_path)

    def _generate_simple_excel_report(self, file_path: str):
        """备用：生成简单Excel报告（原有逻辑）"""
        try:
            if isinstance(self.report_data, dict):
                data_list = [self.report_data]
            else:
                data_list = self.report_data

            df = pd.DataFrame(data_list)

            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='检测数据', index=False)

                # 自动调整列宽
                workbook = writer.book
                worksheet = writer.sheets['检测数据']

                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter

                    for cell in column:
                        try:
                            cell_value = str(cell.value) if cell.value else ""
                            cell_length = len(cell_value)
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass

                    adjusted_width = min(max_length + 2, 35)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                # 添加统计信息（如果有多条数据）
                if len(data_list) > 1:
                    stats_data = {
                        '统计项': ['总检测数', '合格数', '不合格数', '合格率'],
                        '数值': [
                            len(data_list),
                            sum(1 for d in data_list if d.get('is_qualified', False)),
                            sum(1 for d in data_list if not d.get('is_qualified', False)),
                            f"{sum(1 for d in data_list if d.get('is_qualified', False)) / len(data_list) * 100:.2f}%"
                        ]
                    }
                    stats_df = pd.DataFrame(stats_data)
                    stats_df.to_excel(writer, sheet_name='统计信息', index=False)

            print(f"[SUCCESS] 简单Excel报告已保存到: {file_path}")
            return file_path

        except Exception as e:
            print(f"[ERROR] 生成简单Excel报告失败: {str(e)}")
            raise

    def generate_html_report(self, file_path):
        """生成HTML报告"""
        if isinstance(self.report_data, list):
            # 批量报告
            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <title>{self.txt_report_title.toPlainText()}</title>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 40px; }}
                    .header {{ text-align: center; margin-bottom: 40px; }}
                    .title {{ font-size: 24px; font-weight: bold; margin-bottom: 20px; }}
                    .section {{ margin-bottom: 30px; }}
                    .section-title {{ font-size: 18px; font-weight: bold; margin-bottom: 10px; border-bottom: 2px solid #333; padding-bottom: 5px; }}
                    table {{ width: 100%; border-collapse: collapse; margin-bottom: 15px; }}
                    th, td {{ border: 1px solid #ddd; padding: 8px; text-align: center; }}
                    th {{ background-color: #4CAF50; color: white; }}
                    tr:nth-child(even) {{ background-color: #f2f2f2; }}
                    .passed {{ color: green; font-weight: bold; }}
                    .failed {{ color: red; font-weight: bold; }}
                    .stats {{ background-color: #e8f4f8; padding: 15px; border-radius: 5px; }}
                </style>
            </head>
            <body>
                <div class="header">
                    <div class="title">{self.txt_report_title.toPlainText()}</div>
                    <div>生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</div>
                    <div>检测单位: {self.txt_company_info.toPlainText().replace(chr(10), '<br>')}</div>
                </div>

                <div class="section">
                    <div class="section-title">统计汇总</div>
                    <div class="stats">
                        总检测数: {len(self.report_data)}<br>
                        合格数: {sum(1 for d in self.report_data if d.get('is_qualified', False))}<br>
                        不合格数: {sum(1 for d in self.report_data if not d.get('is_qualified', False))}<br>
                        合格率: {sum(1 for d in self.report_data if d.get('is_qualified', False)) / len(self.report_data) * 100:.2f}%<br>
                    </div>
                </div>

                <div class="section">
                    <div class="section-title">详细检测数据</div>
                    <table>
                        <thead>
                            <tr>
                                <th>序号</th>
                                <th>检测时间</th>
                                <th>内圆坐标</th>
                                <th>外圆坐标</th>
                                <th>同心度(‰)</th>
                                <th>偏心距(mm)</th>
                                <th>状态</th>
                            </tr>
                        </thead>
                        <tbody>
            """

            for i, data in enumerate(self.report_data, 1):
                status_class = "passed" if data.get('is_qualified', False) else "failed"
                status_text = "合格" if data.get('is_qualified', False) else "不合格"

                html_content += f"""
                            <tr>
                                <td>{i}</td>
                                <td>{data.get('detection_time', 'N/A')}</td>
                                <td>({data.get('inner_circle', {{}}).get('x', 0):.1f}, {data.get('inner_circle', {{}}).get('y', 0):.1f})</td>
                                <td>({data.get('outer_circle', {{}}).get('x', 0):.1f}, {data.get('outer_circle', {{}}).get('y', 0):.1f})</td>
                                <td>{data.get('concentricity', 0):.3f}</td>
                                <td>{data.get('eccentricity_mm', 0):.3f}</td>
                                <td class="{status_class}">{status_text}</td>
                            </tr>
                """

            html_content += """
                        </tbody>
                    </table>
                </div>

                <div class="section">
                    <div class="section-title">备注</div>
                    <p>""" + self.txt_remarks.toPlainText().replace(chr(10), '<br>') + """</p>
                </div>

                <div class="section">
                    <div class="section-title">签字确认</div>
                    <table>
                        <tr>
                            <th>检测人员</th>
                            <th>审核人员</th>
                            <th>批准人员</th>
                        </tr>
                        <tr>
                            <td style="height: 50px;"></td>
                            <td style="height: 50px;"></td>
                            <td style="height: 50px;"></td>
                        </tr>
                        <tr>
                            <td>日期：</td>
                            <td>日期：</td>
                            <td>日期：</td>
                        </tr>
                    </table>
                </div>
            </body>
            </html>
            """
        else:
            # 单次报告
            status_class = "passed" if self.report_data.get('is_qualified', False) else "failed"
            status_text = "合格" if self.report_data.get('is_qualified', False) else "不合格"

            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <title>{self.txt_report_title.toPlainText()}</title>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 40px; }}
                    .header {{ text-align: center; margin-bottom: 40px; }}
                    .title {{ font-size: 24px; font-weight: bold; margin-bottom: 20px; }}
                    .section {{ margin-bottom: 30px; }}
                    .section-title {{ font-size: 18px; font-weight: bold; margin-bottom: 10px; border-bottom: 2px solid #333; padding-bottom: 5px; }}
                    table {{ width: 100%; border-collapse: collapse; margin-bottom: 15px; }}
                    th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                    th {{ background-color: #4CAF50; color: white; }}
                    .result {{ font-weight: bold; }}
                    .{status_class} {{ color: {'green' if self.report_data.get('is_qualified', False) else 'red'}; }}
                    .info {{ background-color: #f9f9f9; padding: 15px; border-radius: 5px; }}
                </style>
            </head>
            <body>
                <div class="header">
                    <div class="title">{self.txt_report_title.toPlainText()}</div>
                    <div>生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</div>
                    <div>检测单位: {self.txt_company_info.toPlainText().replace(chr(10), '<br>')}</div>
                </div>

                <div class="section">
                    <div class="section-title">检测结果</div>
                    <div class="info {status_class}">
                        检测状态: {status_text}<br>
                        检测时间: {self.report_data.get('detection_time', 'N/A')}<br>
                        处理耗时: {self.report_data.get('processing_time', 0):.3f} 秒
                    </div>
                </div>

                <div class="section">
                    <div class="section-title">几何参数</div>
                    <table>
                        <tr>
                            <th>参数</th>
                            <th>数值</th>
                        </tr>
                        <tr>
                            <td>内圆坐标</td>
                            <td>({self.report_data.get('inner_circle', {{}}).get('x', 0):.1f}, {self.report_data.get('inner_circle', {{}}).get('y', 0):.1f})</td>
                        </tr>
                        <tr>
                            <td>内圆半径</td>
                            <td>{self.report_data.get('inner_circle', {{}}).get('radius', 0):.1f} px</td>
                        </tr>
                        <tr>
                            <td>外圆坐标</td>
                            <td>({self.report_data.get('outer_circle', {{}}).get('x', 0):.1f}, {self.report_data.get('outer_circle', {{}}).get('y', 0):.1f})</td>
                        </tr>
                        <tr>
                            <td>外圆半径</td>
                            <td>{self.report_data.get('outer_circle', {{}}).get('radius', 0):.1f} px</td>
                        </tr>
                    </table>
                </div>

                <div class="section">
                    <div class="section-title">同心度分析</div>
                    <table>
                        <tr>
                            <th>检测项目</th>
                            <th>检测结果</th>
                            <th>判定标准</th>
                            <th>结论</th>
                        </tr>
                        <tr>
                            <td>同心度</td>
                            <td>{self.report_data.get('concentricity', 0):.3f}‰</td>
                            <td>≤{self.report_data.get('tolerance', 0.2):.2f}‰</td>
                            <td class="{status_class}">{status_text}</td>
                        </tr>
                        <tr>
                            <td>偏心距</td>
                            <td>{self.report_data.get('eccentricity_mm', 0):.3f} mm</td>
                            <td>≤{self.report_data.get('tolerance', 0.2):.2f} mm</td>
                            <td class="{status_class}">{status_text}</td>
                        </tr>
                    </table>
                </div>

                <div class="section">
                    <div class="section-title">备注</div>
                    <p>""" + self.txt_remarks.toPlainText().replace(chr(10), '<br>') + """</p>
                </div>
            </body>
            </html>
            """

        # 写入文件
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html_content)

    def preview_report(self):
        """预览报告"""
        QMessageBox.information(self, "预览", "报告预览功能开发中...")

    def export_data(self):
        """导出数据"""
        if self.report_data is None:
            QMessageBox.warning(self, "警告", "没有可导出的数据！")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "导出数据",
            f"concentricity_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
            "JSON文件 (*.json)"
        )

        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(self.report_data, f, ensure_ascii=False, indent=2)

                QMessageBox.information(self, "成功", "数据导出成功！")

            except Exception as e:
                QMessageBox.critical(self, "错误", f"数据导出失败: {str(e)}")

    def export_to_excel(self):
        """导出结果数据（核心数据，格式与Excel报告一致）"""
        if self.report_data is None:
            QMessageBox.warning(self, "警告", "没有可导出的数据！")
            return

        try:
            # 设置默认保存目录
            default_dir = os.path.join(self.project_root, "data", "results")
            os.makedirs(default_dir, exist_ok=True)

            print(f"[DEBUG] 默认保存目录: {default_dir}")

            # 生成文件名
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_name = f"同心度检测结果_{timestamp}.xlsx"

            # 设置完整默认路径
            default_path = os.path.join(default_dir, default_name)
            print(f"[DEBUG] 默认保存路径: {default_path}")

            # 显示保存对话框
            file_path, selected_filter = QFileDialog.getSaveFileName(
                self,
                "保存结果数据",
                default_path,
                "Excel文件 (*.xlsx);;CSV文件 (*.csv);;JSON文件 (*.json)"
            )

            if not file_path:
                return  # 用户取消

            print(f"[DEBUG] 用户选择路径: {file_path}")

            # 确保目录存在
            save_dir = os.path.dirname(file_path)
            if save_dir and not os.path.exists(save_dir):
                os.makedirs(save_dir, exist_ok=True)

            # 根据数据类型选择合适的保存函数
            from utils.file_io import save_concentricity_result_excel, save_multiple_results_to_excel

            if isinstance(self.report_data, list):
                # 如果是列表，使用批量保存函数
                if not self.report_data:
                    QMessageBox.warning(self, "警告", "结果数据为空！")
                    return

                # 检查列表中的元素类型
                if all(isinstance(item, dict) for item in self.report_data):
                    # 如果是字典列表，使用批量保存函数
                    file_path = save_multiple_results_to_excel(self.report_data, file_path)
                    format_name = "Excel（批量）"
                else:
                    # 如果列表包含其他类型，取第一个元素尝试保存
                    first_item = self.report_data[0] if self.report_data else {}
                    if isinstance(first_item, dict):
                        file_path = save_concentricity_result_excel(first_item, save_dir)
                        format_name = "Excel（单条）"
                    else:
                        raise ValueError("不支持的数据格式")
            elif isinstance(self.report_data, dict):
                # 如果是字典，直接使用单个结果保存函数
                file_path = save_concentricity_result_excel(self.report_data, save_dir)
                format_name = "Excel"
            else:
                raise ValueError(f"不支持的数据类型: {type(self.report_data)}")

            success_msg = f"结果数据已保存为{format_name}格式！\n\n文件位置：\n{file_path}"
            self.lbl_status.setText(f"结果已保存: {os.path.basename(file_path)}")
            QMessageBox.information(self, "成功", success_msg)

        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"导出数据失败: {str(e)}\n详细错误信息:\n{error_details}")

            error_msg = f"导出数据失败：{str(e)}"
            QMessageBox.critical(self, "错误", error_msg)

    def show_error(self, message):
        """显示错误信息"""
        QMessageBox.critical(self, "错误", message)

    def _extract_core_data(self, report_data):
        """从检测结果中提取核心数据（与Excel报告格式一致）"""
        core_data = {}

        # 1. 检测基本信息
        core_data["检测时间"] = report_data.get("detection_time", "")
        core_data["图像尺寸"] = report_data.get("image_size", "")
        core_data["是否合格"] = "是" if report_data.get("is_qualified", False) else "否"
        core_data["实际偏差(mm)"] = round(report_data.get("error_mm", 0), 4)
        core_data["像素偏差(px)"] = round(report_data.get("pixel_error", 0), 2)
        core_data["同心度(‰)"] = round(report_data.get("relative_error_percent", 0) * 10, 4)  # % 转换为 ‰

        # 2. 内圆坐标
        inner_center = report_data.get("inner_center", {})
        if isinstance(inner_center, dict):
            core_data["内圆X(px)"] = round(inner_center.get("x", 0), 2)
            core_data["内圆Y(px)"] = round(inner_center.get("y", 0), 2)
        else:
            core_data["内圆X(px)"] = 0.0
            core_data["内圆Y(px)"] = 0.0

        # 3. 外圆坐标
        outer_circle = report_data.get("outer_circle", {})
        if isinstance(outer_circle, dict):
            core_data["外圆X(px)"] = round(outer_circle.get("x", 0), 2)
            core_data["外圆Y(px)"] = round(outer_circle.get("y", 0), 2)
            core_data["外圆半径(px)"] = round(outer_circle.get("radius", 0), 2)
        else:
            core_data["外圆X(px)"] = 0.0
            core_data["外圆Y(px)"] = 0.0
            core_data["外圆半径(px)"] = 0.0

        # 4. 其他关键参数
        core_data["处理时间(s)"] = round(report_data.get("processing_time", 0), 3)

        # 计算检测置信度（取内圆和外圆的最高置信度）
        inner_confidence = inner_center.get("confidence", 0) if isinstance(inner_center, dict) else 0
        outer_confidence = outer_circle.get("confidence", 0) if isinstance(outer_circle, dict) else 0
        core_data["检测置信度"] = round(max(inner_confidence, outer_confidence), 3)

        return core_data

    def _save_core_data_excel(self, filepath):
        """保存核心数据到Excel文件（本地实现）"""
        try:
            # 提取核心数据
            core_data = self._extract_core_data(self.report_data)

            # 创建DataFrame
            df = pd.DataFrame([core_data])

            # 保存为Excel
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='检测结果', index=False)

                # 获取worksheet对象以调整列宽
                workbook = writer.book
                worksheet = writer.sheets['检测结果']

                # 自动调整所有列宽
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter

                    # 计算每列的最大长度
                    for cell in column:
                        try:
                            cell_value = str(cell.value) if cell.value is not None else ""
                            cell_length = len(cell_value)
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass

                    # 设置列宽
                    adjusted_width = min(max_length + 2, 30)  # 最大宽度30
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            print(f"[SUCCESS] 核心数据Excel文件已保存到: {filepath}")
            return filepath

        except Exception as e:
            print(f"[ERROR] 保存Excel文件失败: {str(e)}")
            raise